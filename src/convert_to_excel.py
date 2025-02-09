import json
import shutil
import openpyxl

json_path = "./psopk.com.json"
xlsx_template = "./Template.xlsx"
xlsx_output = "./Output.xlsx"

shutil.copy(xlsx_template, xlsx_output)

with open(json_path, "r", encoding="utf-8") as file:
    data = json.load(file)

wb = openpyxl.load_workbook(xlsx_output)
ws = wb.active

emails = data.get("emails", [])
phone_numbers = data.get("phone_numbers", [])
social_media = data.get("social_media", [])
stats = data.get("stats", {})

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
    for cell in row:
        cell.value = None

for i, email in enumerate(emails, start=2):
    ws[f"A{i}"] = email

for i, phone in enumerate(phone_numbers, start=2):
    ws[f"B{i}"] = phone

for i, link in enumerate(social_media, start=2):
    ws[f"C{i}"] = link

ws["F2"] = stats.get("EXTRACTED_EMAILS", 0)
ws["F3"] = stats.get("EXTRACTED_PHONE_NUMBERS", 0)
ws["F4"] = stats.get("TOTAL_SOCIAL_MEDIA_LINKS", 0)

wb.save(xlsx_output)
